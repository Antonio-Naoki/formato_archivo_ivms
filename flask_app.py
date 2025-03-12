from flask import Flask, request, render_template, send_file, redirect, url_for
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import zipfile
import io

app = Flask(__name__)

# función para aplicar estilos a la tabla
def aplicar_estilos(ws):
    # estilo para los encabezados
    header_fill = PatternFill(start_color="3CB371", end_color="3CB371", fill_type="solid") 
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # estilo para las celdas de datos
    data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Arial", size=11, color="000000")
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    data_alignment = Alignment(horizontal="center", vertical="center")

    # aplicar estilos a los encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment

    # aplicar estilos a las celdas de datos
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = data_fill
            cell.font = data_font
            cell.border = data_border
            cell.alignment = data_alignment

    # ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

# función para procesar un archivo
def procesar_archivo(df, nombre_archivo):
    # eliminar columnas no deseadas
    columnas_a_eliminar = ["Data Source", "Handling Type", "Temperature", "Abnormal", "Attendance Check Point"]
    df = df.drop(columns=columnas_a_eliminar, errors='ignore')

    # cambiar nombres de columnas a español
    df = df.rename(columns={
        "Person ID": "ID Persona",
        "Name": "Nombre",
        "Department": "Departmento",
        "Time": "Hora",
        "Attendance Status": "Estado de Asistencia",
        "Custom Name": "Tipo de Evento"
    })

    # formatear la columna "Hora" a formato de hora
    df["Hora"] = pd.to_datetime(df["Hora"]).dt.strftime('%Y-%m-%d %H:%M:%S')

    # guardar el archivo formateado
    nombre_salida = f"{nombre_archivo}_formateado.xlsx" # para un servidor como pythonanywhere.com hay que usar esta ruta para los archivos de descarga: f"/tmp/{nombre_archivo}_formateado.xlsx"
    # para vercel debo probarprimero.
    
    # crear un nuevo archivo Excel con openpyxl
    wb = Workbook()
    ws = wb.active

    # escribir los encabezados
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # escribir los datos
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # aplicar estilos a la tabla
    aplicar_estilos(ws)

    # guardar el archivo
    wb.save(nombre_salida)

    return nombre_salida

# ruta principal que muestra el formulario
@app.route('/')
def index():
    return render_template('index.html')

# ruta para manejar la carga de múltiples archivos y procesamiento
@app.route('/procesar', methods=['POST'])
def procesar_archivos_route():
    if 'files' not in request.files:
        return "No se han subido archivos"
    
    archivos = request.files.getlist('files')

    if not archivos or all(archivo.filename == '' for archivo in archivos):
        return "No se han seleccionado archivos"

    # crear un archivo ZIP en memoria para almacenar los archivos procesados
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for archivo in archivos:
            if archivo.filename.endswith('.csv'):
                df = pd.read_csv(archivo)
            elif archivo.filename.endswith('.xlsx'):
                df = pd.read_excel(archivo)
            else:
                continue  # saltar archivos no soportados

            # obtener el nombre del archivo sin la extensión
            nombre_archivo = os.path.splitext(archivo.filename)[0]

            # procesar el archivo
            archivo_salida = procesar_archivo(df, nombre_archivo)

            # agregar el archivo procesado al ZIP
            zip_file.write(archivo_salida, os.path.basename(archivo_salida))

            # eliminar el archivo temporal
            os.remove(archivo_salida)

    # preparar el archivo ZIP para descarga
    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='archivos_procesados.zip'
    )

if __name__ == '__main__':
    app.run(debug=True)


# from flask import Flask, request, render_template, send_file
# import pandas as pd
# from datetime import datetime
# import os
# import zipfile
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# app = Flask(__name__)
# UPLOAD_FOLDER = os.path.abspath("uploads")
# PROCESSED_FOLDER = os.path.abspath("processed")
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(PROCESSED_FOLDER, exist_ok=True)

# # Función para procesar el archivo
# def procesar_archivo(df, nombre_archivo):
#     columnas_a_eliminar = ["Data Source", "Handling Type", "Temperature", "Abnormal", "Attendance Check Point"]
#     df = df.drop(columns=columnas_a_eliminar, errors='ignore')
    
#     df = df.rename(columns={
#         "Person ID": "ID Persona",
#         "Name": "Nombre",
#         "Time": "Hora",
#         "Attendance Status": "Estado de Asistencia",
#         "Custom Name": "Tipo de Evento"
#     })
    
#     if "Hora" in df.columns:
#         df["Hora"] = pd.to_datetime(df["Hora"], errors='coerce')
#         df = df.sort_values(by="Hora", ascending=True)  # Orden ascendente (de más temprano a más tarde)
#         df["Hora"] = df["Hora"].dt.strftime('%I:%M %p')  # Formato de hora normal sin fecha
    
#     if "Tipo de Evento" in df.columns:
#         df["Tipo de Evento"] = df["Tipo de Evento"].replace({"Entrada": "Entrada", "Salida": "Salida"})
    
#     nombre_salida = os.path.join(PROCESSED_FOLDER, f"{nombre_archivo}_formateado.xlsx")
    
#     with pd.ExcelWriter(nombre_salida, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name="Datos")
#         workbook = writer.book
#         worksheet = writer.sheets["Datos"]
        
#         # Estilos para encabezados
#         header_fill = PatternFill(start_color="3CB371", end_color="3CB371", fill_type="solid")  # Verde profesional
#         header_font = Font(color="FFFFFF", bold=True, size=12)
#         header_alignment = Alignment(horizontal="center", vertical="center")
#         thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
#         for cell in worksheet[1]:
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = header_alignment
#             cell.border = thin_border
        
#         # Estilos para filas (Fondo blanco, bordes definidos)
#         for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
#             for cell in row:
#                 cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
#                 cell.border = thin_border
#                 cell.alignment = Alignment(horizontal="left", vertical="center")
        
#         # Ajustar el tamaño de las columnas automáticamente
#         for col in worksheet.columns:
#             max_length = 0
#             col_letter = col[0].column_letter
#             for cell in col:
#                 try:
#                     if cell.value:
#                         max_length = max(max_length, len(str(cell.value)))
#                 except:
#                     pass
#             adjusted_width = (max_length + 2)
#             worksheet.column_dimensions[col_letter].width = adjusted_width
    
#     return nombre_salida

# @app.route('/')
# def index():
#     return render_template('index.html')

# @app.route('/procesar', methods=['POST'])
# def procesar_archivos_route():
#     if 'files' not in request.files:
#         return "No se han subido archivos"
    
#     archivos = request.files.getlist('files')
#     if not archivos or all(archivo.filename == '' for archivo in archivos):
#         return "No se han seleccionado archivos"
    
#     archivos_procesados = []
    
#     for archivo in archivos:
#         nombre_archivo = os.path.splitext(archivo.filename)[0]
#         ruta_archivo = os.path.join(UPLOAD_FOLDER, archivo.filename)
#         archivo.save(ruta_archivo)
        
#         try:
#             if archivo.filename.endswith('.csv'):
#                 df = pd.read_csv(ruta_archivo)
#             elif archivo.filename.endswith('.xlsx'):
#                 df = pd.read_excel(ruta_archivo, engine='openpyxl')
#             else:
#                 continue
            
#             archivo_salida = procesar_archivo(df, nombre_archivo)
#             archivos_procesados.append(archivo_salida)
#         except Exception as e:
#             continue
    
#     if not archivos_procesados:
#         return "No se pudieron procesar los archivos"
    
#     zip_salida = os.path.join(PROCESSED_FOLDER, "archivos_procesados.zip")
#     with zipfile.ZipFile(zip_salida, 'w') as zipf:
#         for archivo in archivos_procesados:
#             zipf.write(archivo, os.path.basename(archivo))
    
#     return send_file(zip_salida, as_attachment=True, mimetype='application/zip')

# if __name__ == '__main__':
#     app.run(debug=True)
